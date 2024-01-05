import pandas as pd
import openpyxl as py
import os
import re
from openpyxl.utils.dataframe import dataframe_to_rows


def excel_ref_to_indices(ref):
    # Regular expression pattern to split Excel-like reference strings
    pattern = r'([A-Z]+)(\d+):([A-Z]+)(\d+)'

    # Match the pattern in the reference string
    match = re.match(pattern, ref)

    if match:
        start_col, start_row, end_col, end_row = match.groups()

        # Function to convert column reference to index
        def col_ref_to_index(col_ref):
            num = 0
            for i, c in enumerate(reversed(col_ref)):
                num += (ord(c) - ord('A') + 1) * (26 ** i)
            return num

        # Convert references to indices
        start_col = col_ref_to_index(start_col)
        end_col = col_ref_to_index(end_col)
        start_row = int(start_row)
        end_row = int(end_row)

        return start_col, start_row, end_col, end_row
    else:
        return None


def get_df_from_cell_reference(sheet, ref, index=True, header=True):
    """
    Reads data from a specified range reference in an Excel file and returns it as a DataFrame.

    Parameters:
    - sheet_name: Name of the Excel sheet.
    - ref: Cell range reference (e.g., 'A1:D10').

    Returns:
    - DataFrame containing data from the specified range reference.
    """
    try:
        # Parsing reference string to get start and end cells
        start_col, start_row, end_col,  end_row = excel_ref_to_indices(ref)
        first_row = True
        index_values = []
        # Accessing data within the specified range
        data = []
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row,
                                   min_col=start_col, max_col=end_col,
                                   values_only=True
                                   ):
            if first_row:
                header_values = row[1:]
            else:
                data.append(row[1:])
                index_values.append(row[0])
            first_row = False

        # Creating a DataFrame
        df = pd.DataFrame(data, index=index_values, columns=header_values)
        return df

    except Exception as e:
        print(f"Error reading Excel file: {e}")
        raise


def insert_dataframes_as_tables(list_of_dicts, excel_filename='output.xlsx', sheet_name='Sheet1', orientation='vertical'):
    wb = py.Workbook()
    ws = wb.active
    ws.title = sheet_name

    row_counter = 1
    col_counter = 1

    # Check if the input is a list of dictionaries
    keys = set().union(*list_of_dicts)
    for key in keys:
        for df_dict in list_of_dicts:
            df = df_dict.get(key)
            if df is not None:
                rows = dataframe_to_rows(df, index=False, header=True)
                for r_idx, row in enumerate(rows, 1):
                    if orientation == 'vertical':
                        for c_idx, value in enumerate(row, 1):
                            ws.cell(row=row_counter + r_idx, column=col_counter, value=value)
                    elif orientation == 'horizontal':
                        for c_idx, value in enumerate(row, 1):
                            ws.cell(row=row_counter, column=col_counter + c_idx, value=value)

                if orientation == 'vertical':
                    row_counter += df.shape[0] + 2  # Increase row counter for the next table and add some space
                elif orientation == 'horizontal':
                    col_counter += df.shape[1] + 2  # Increase column counter for the next table and add some space
            else:
                if orientation == 'vertical':
                    row_counter += 2  # If a key is missing in any dictionary, leave space for it
                elif orientation == 'horizontal':
                    col_counter += 2  # If a key is missing in any dictionary, leave space for it

        if orientation == 'vertical':
            row_counter = 1  # Reset row counter for the next set of dataframes with different keys
            col_counter += 2  # Add space between sets of dataframes
        elif orientation == 'horizontal':
            col_counter = 1  # Reset col counter for the next set of dataframes with different keys
            row_counter += 2  # Add space between sets of dataframes

    wb.save(filename=excel_filename)
    print(f"Dataframes inserted as tables to '{excel_filename}' ({orientation} orientation) successfully.")



if __name__ == "__main__":
    def get_map_df():
        file_path = "../data/T400_5N_dataset_step120.xlsx"
        wb = py.load_workbook(filename=file_path, data_only=True)
        ws = wb["MAPS"]
        map_table = ws.tables["KFMSWDKQ"]
        ref = map_table.ref
        map_df = get_df_from_cell_reference(ws, map_table.ref)
        try:
            map_df.columns = map_df.columns.astype("float")
        except:
            pass
        map_df.reset_index(inplace=True)
        df_long = map_df.melt(id_vars='index', var_name='y', value_name='value')
        df_long.columns = ["x", "y", "value"]
        return map_df

    get_map_df()