import pandas as pd
from openpyxl import load_workbook
import re


def read_excel_range_reference(sheet, ref, index=True, header=True):
    """
    Reads data from a specified range reference in an Excel file and returns it as a DataFrame.

    Parameters:
    - file_path: Path to the Excel file.
    - sheet_name: Name of the Excel sheet.
    - ref: Cell range reference (e.g., 'A1:D10').

    Returns:
    - DataFrame containing data from the specified range reference.
    """
    try:
        # Parsing reference string to get start and end cells
        start_col, start_row, end_col,  end_row = re.findall(r'[A-Z]+|\d+', ref)
        start_col = ord(start_col) - ord('A') + 1
        start_row = int(start_row)
        end_col = ord(end_col) - ord('A') + 1
        end_row = int(end_row)
        first_row = True
        index_values = []
        # Accessing data within the specified range
        data = []
        for row in sheet.iter_rows(min_row=start_row, max_row=end_row,
                                   min_col=start_col, max_col=end_col,
                                   values_only=True
                                   ):
            if first_row:
                header_values = row[1:]
            else:
                data.append(row[1:])
                index_values.append(row[0])
            first_row = False

        # Creating a DataFrame
        df = pd.DataFrame(data, index=index_values, columns=header_values)
        return df
    except Exception as e:
        raise
        print(f"Error reading Excel file: {e}")
        return None


file_path = "../../data/T400_5N_dataset_step120.xlsx"
wb = load_workbook(filename=file_path, data_only=True)
ws = wb["MAPS"]
map_table = ws.tables["KFMSWDKQ"]
ref = map_table.ref
map_df = read_excel_range_reference(ws, map_table.ref)